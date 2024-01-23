import pyautogui as pg
import time
import os
from openpyxl import load_workbook
import clipboard as cb

timeBetweenPages = pg.prompt('Tempo entre paginas:')
timeBetweenFields = pg.prompt('Tempo entre campos:')

workbook = load_workbook(filename="C:\\laragon\www\\rpa_cad_functionarios_dominio\\base.xlsx")

# Selecionar a aba ativa
sheet = workbook.active

def copyCell(cell):
    cb.copy(str(sheet[cell+str(i)].value).replace("'",""))
    time.sleep(.5)
    if(str(sheet[cell+str(i)].value).replace("'","") != ''):
        pg.hotkey('ctrl', 'v')
    time.sleep(.5)

def copyCellMoeda(cell):
    valor = str(sheet[cell+str(i)].value).replace("'","")
    valor = valor.replace('.', ',')
    time.sleep(.5)
    if(str(sheet[cell+str(i)].value).replace("'","") != ''):
        pg.write(valor,0.2)
    time.sleep(.5)

def pressCell(cell):
    for letra in str(sheet[cell+str(i)].value).replace("'",""):
        if(letra == ','):
            pg.hotkey('right')
        else:
            pg.hotkey(letra)
            time.sleep(.5)

def iniciar():
    pg.keyDown('win')
    pg.hotkey('1')
    pg.hotkey('1')
    pg.keyUp('win')

def geral():
    time.sleep(float(timeBetweenFields))
    pg.hotkey('alt')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('a')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('e')
    time.sleep(float(timeBetweenFields))
    time.sleep(float(timeBetweenPages))
    pg.hotkey('enter')
    time.sleep(3)
    pg.hotkey('tab')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('delete')
    copyCell('A') #código esocial
    pg.hotkey('tab')
    time.sleep(float(timeBetweenFields))
    copyCell('B') #Nome funcionário
    pg.press('tab', 2)
    copyCell('C') #CPF
    pg.hotkey('tab')
    time.sleep(float(timeBetweenFields))
    copyCell('D') #PIS
    pg.press('tab', 2)
    copyCell('E') #Serviço
    pg.hotkey('tab')
    time.sleep(float(timeBetweenFields))
    copyCell('F') #cargo
    pg.hotkey('tab')
    time.sleep(float(timeBetweenFields))
    copyCell('G') #função
    pg.hotkey('tab')
    time.sleep(float(timeBetweenFields))
    copyCell('H') #departamento
    pg.hotkey('tab')
    time.sleep(float(timeBetweenFields))
    copyCell('I') #centro de custo
    pg.hotkey('tab')
    time.sleep(float(timeBetweenFields))
    copyCell('J') #sindicato
    pg.hotkey('tab')
    time.sleep(float(timeBetweenFields))
    copyCell('K') #admissão
    pg.press('tab', 10)
    time.sleep(float(timeBetweenFields))
    pg.press('backspace', 7)
    time.sleep(float(timeBetweenFields))
    copyCellMoeda('L') #Salário
    pg.press('tab', 10)
    pg.hotkey('right')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 20)
    pg.hotkey('right')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 8)
    time.sleep(float(timeBetweenFields))
    copyCell('M') #CTPS Numero
    pg.press('tab')
    time.sleep(float(timeBetweenFields))
    copyCell('N') #CTPS Serie
    pg.press('tab')
    time.sleep(float(timeBetweenFields))
    copyCell('O') #CTPS Expedição
    time.sleep(float(timeBetweenFields))
    pg.press('tab')
    pressCell('P') #UF Cart. Prof
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 14)
    time.sleep(float(timeBetweenFields))
    pg.press('right')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 2)
    copyCell('Q') #CTPS Expedição
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 10)
    pressCell('R') #
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('S') #
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('T') #
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('U') #
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    pressCell('V') #
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 16)
    pg.hotkey('right')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 2)
    copyCell('W')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 12)
    pg.hotkey('right')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 3)
    copyCell('X')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    pressCell('Y')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('Z')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    pressCell('AA')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AB')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 17)
    pg.hotkey('right')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 6)
    copyCell('AC')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AD')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AE')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 8)
    copyCell('AF')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 12)
    time.sleep(float(timeBetweenFields))
    pg.press('right', 2)
    pg.press('tab', 2)
    copyCell('AG')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    pressCell('AH')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 2)
    copyCell('AI')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AJ')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AK')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AL')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 3)
    copyCell('AM')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 2)
    copyCell('AN')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AO')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AP')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AQ')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AR')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AS')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 15)
    time.sleep(float(timeBetweenFields))
    pg.hotkey('right')
    pg.press('tab', 2)
    copyCell('AT')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AU')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AV')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 2)
    copyCell('AW')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    copyCell('AX')
    time.sleep(float(timeBetweenFields))
    pg.press('tab', 12)
    time.sleep(float(timeBetweenFields))
    pg.hotkey('right')
    pg.press('tab', 4)
    pressCell('AY')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    pressCell('AZ')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    pressCell('BA')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('tab')
    pressCell('BB')
    time.sleep(float(timeBetweenFields))
    pg.hotkey('alt', 'G')
    time.sleep(float(timeBetweenPages))
    pg.hotkey('n')
    time.sleep(float(timeBetweenPages))
    pg.hotkey('s')
    time.sleep(5)
    pg.hotkey('esc')
    time.sleep(5)





# Iterar pelas linhas
iniciar()
for i in range(2, sheet.max_row):
    geral()

pg.alert('Bot Finalizado')
    

